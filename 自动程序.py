print("welcome！！欢迎使用该程序！！")

import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment

# 获取输入的表格名称
oigin = input("请输入志愿者名单表格的名称：")

# 读取前5行数据
df = pd.read_excel(f"{oigin}.xlsx", header=None, nrows=5)

# 找到包含"序号"列标题的行
header_row = None
for i in range(5):
    if "序号" in df.iloc[i].values:
        header_row = i
        print(f"找到列标题在第{i+1}行")
        break

# 如果找到列标题行，重新读取整个表格并指定列标题
if header_row is not None:
    df = pd.read_excel(f"{oigin}.xlsx", header=header_row)

    # 去除列标题的前后空格
    df.columns = df.columns.str.strip()

    # 调整列的顺序
    new_order = ['序号', '学号', '姓名', '书院班级', '专业班级', '联系方式', '所属书院']

    # 检查列名是否完全匹配
    if set(new_order).issubset(df.columns):
        df = df[new_order]
        # 保存调整后的表格
        df.to_excel('调整后的表格.xlsx', index=False)
        print("表格调整完成并保存为 '调整后的表格.xlsx'")
    else:
        print("列名不匹配，可能是有些列标题没有与pu表中的格式对应。请检查列标题是否正确。")
        exit()
else:
    print("没有找到'序号'列标题，可能是有些列标题没有与pu表中的格式对应")
    exit()

# 读取表格
df = pd.read_excel('调整后的表格.xlsx')

# 获取最后一行的行号（索引位置）
last_row_index = df.index[-1] + 2

# 坐标字母转数字函数
def column_to_number(col_str):
    col_num = 0
    for char in col_str:
        col_num = col_num * 26 + ord(char.upper()) - ord('A') + 1
    return col_num

# 坐标数字转字母函数
def number_to_column(col_num):
    col_str = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        col_str = chr(remainder + ord('A')) + col_str
    return col_str

# 用户输入表格标题、活动信息
title = input("请输入表格标题：")
event_date = input("请输入活动时间（例如：2025年01月15日）：")
event_duration = input("请输入活动时长（小时）：")
event_points = input("请输入志愿积分：")
event_contact = input("请输入活动联系人：")
event_description = input("请输入活动介绍：")

# coord1 固定为 A2
coord1 = 'A2'  # 固定为 A2

# coord2 的横坐标固定为 G，纵坐标为 last_row_index 的值
coord2 = f"G{last_row_index}"

coord1_col, coord1_row = ord(coord1[0].upper()) - ord('A'), int(coord1[1:])
coord2_col, coord2_row = ord(coord2[0].upper()) - ord('A'), int(coord2[1:])
num_people = abs(coord2_row - coord1_row) + 1
coordC = coord2_row
coord3 = 2 + num_people
coord4 = 5 + num_people
coord5 = 4 + 2 * num_people
coordA = 4 + num_people
coordB = 4 + num_people
coordI = 4 + num_people
coordK = 4 + 2 * num_people
coordH = 2 + num_people

# 解析坐标1和坐标2
start_col_str, start_row = coord1[0], int(coord1[1:])
end_col_str, end_row = coord2[0], int(coord2[1:])
start_col = column_to_number(start_col_str)  # 将列字母转换为列数字
end_col = column_to_number(end_col_str)      # 将列字母转换为列数字

# 计算表格B的坐标范围和人数
num_rows = abs(end_row - start_row) + 1  # 计算人数

# 创建表格X
wb_x = openpyxl.Workbook()
ws_x = wb_x.active
ws_x.title = "sheet1"

# 在单元格A1输入内容并设置样式
ws_x["A1"] = f"{title}志愿活动志愿者名单"
ws_x.merge_cells("A1:H1")
ws_x["A1"].font = Font(size=16, bold=True)
ws_x["A1"].alignment = Alignment(horizontal="center", vertical="center")

# 在A2到D2输入内容并设置样式
headers = ["序号", "学号", "姓名", "备注"]
for col, header in enumerate(headers, start=1):
    cell = ws_x.cell(row=2, column=col)
    cell.value = header
    cell.font = Font(size=12, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 在F2单元格输入内容并设置样式
ws_x["F2"] = "←（官Q公示名单）"
ws_x["F2"].font = Font(size=14, bold=True, color="FF0000")
ws_x["F2"].alignment = Alignment(horizontal="center", vertical="center")
ws_x.merge_cells(f"F2:H{coordH}")

# 在coordA输入内容并设置样式
extended_headers = ["序号", "学号", "姓名", "书院班级", "专业班级", "联系方式", "所属书院", "备注"]
for col, header in enumerate(extended_headers, start=1):
    cell = ws_x.cell(row=coordA, column=col)
    cell.value = header
    cell.font = Font(size=12, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 在I22单元格输入内容并设置样式
ws_x[f"I{coordI}"] = "←（PU正式导入名单）"
ws_x[f"I{coordI}"].font = Font(size=14, bold=True, color="FF0000")
ws_x[f"I{coordI}"].alignment = Alignment(horizontal="center", vertical="center")
ws_x.merge_cells(f"I{coordI}:K{coordK}")

# 设置列宽
column_widths = {"B": 12, "C": 10, "D": 10, "E": 20, "F": 15, "G": 11}
for col, width in column_widths.items():
    ws_x.column_dimensions[col].width = width

# 打开调整后的表格
wb_b = openpyxl.load_workbook("调整后的表格.xlsx")
ws_b = wb_b.active

# 复制调整后的表格的内容到表格X
for row_idx, row in enumerate(ws_b.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=3), start=3):
    for col_idx, cell in enumerate(row, start=1):
        ws_x.cell(row=row_idx, column=col_idx).value = cell.value
        ws_x.cell(row=row_idx, column=col_idx).font = Font(size=11)
        ws_x.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")

for row_idx, row in enumerate(ws_b.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=7), start=coord4):
    for col_idx, cell in enumerate(row, start=1):  
        ws_x.cell(row=row_idx, column=col_idx).value = cell.value
        ws_x.cell(row=row_idx, column=col_idx).font = Font(size=11)
        ws_x.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")

# 在表格A末尾添加备注
start_row = num_rows + num_rows + 6
ws_x[f"A{start_row}"] = "备注："
ws_x[f"A{start_row}"].font = Font(size=11)

# 添加活动内容
activity_rows = [
    f"1.活动时间：{event_date}",
    "2.活动设置为后台签到",
    f"3.活动时长{event_duration}小时，志愿积分{event_points}积分",
    f"4.活动联系人：{event_contact}",
    f"5.活动介绍：{event_description}"
]
for i, content in enumerate(activity_rows, start=start_row + 1):
    ws_x[f"A{i}"] = content
    ws_x.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    ws_x[f"A{i}"].font = Font(size=11)
    ws_x[f"A{i}"].alignment = Alignment(horizontal="left", vertical="center")

# 保存表格X
wb_x.save(f"{title}.xlsx")
print(f"表格{title}已创建完成！")
